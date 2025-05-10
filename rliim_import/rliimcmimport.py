# ***********************************************************
# RLIIM CM IMPORT
#
# Import of collected material from the RLIIM lot spreadsheet to Kiosk
#
# (c) Lutz Klein 2025
#
# ***********************************************************

import logging
import re
import sys
import warnings
from os import path
import datetime
from typing import List

import psycopg2
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

import kioskstdlib
from kiosksqldb import KioskSQLDb
from sync_config import SyncConfig
from sync_plugins.rliim_kiosk_bridge.rliim_import.rliim2kioskimport import import_rliim_cm_for_new_loci

ESC_RED = "\u001b[31m"
ESC_GREEN = "\u001b[32;1m"
ESC_YELLOW = "\u001b[33;1m"
ESC_RESET = "\u001b[0m"

params = {"-db": "db",
          "--db": "db",
          "-d": "dev",
          "--dev": "dev",
          "--no_context_check": "no_context_check",
          "-nc": "no_context_check",
          "--import": "import",
          "-i": "import",
          "--user": "user",
          "--log_file": "logfile",
          "-lf": "logfile",
          "--no_commit": "nc",
          "--erase_former_import": "efi",
          "--apply_imports": "apply"
          }

cfg: SyncConfig
kiosk_dir = ""

context_register = {}

added_identifier_log = []
added_identifiers = []

options = {}
dsd_file = ""
database = ""
workbook_file = ""
RLIIM_IMPORT_TABLE = "rliim_cm_import"

expected_cols_lots = ["lot number", "context", "trench", "type", "date excavated", "extra info", "date uploaded",
                      "total number of sherds",
                      "recorded",
                      "photographed"]

expected_cols_samples = ["Sample Number", "context", "trench", "type", "description", "date", "notes", "location"]

expected_cols_artifacts = ["lot number", "context", "trench", "type", "date excavated", "other dimensions", "weight",
                           "description", "date entered", "photographed?"]

def connect_database():
    con = psycopg2.connect(f"dbname={cfg.database_name} user={cfg.database_usr_name} password={cfg.database_usr_pwd}")
    con.autocommit = False
    KioskSQLDb._con_ = con
    cur = KioskSQLDb.execute_return_cursor("select current_database()")
    try:
        r = cur.fetchone()
        assert r[0] == cfg.database_name
    finally:
        cur.close()


# **********************************************
# console part
# **********************************************

SEPARATOR = -2
INFO = -1
SUCCESS = 0
WARNING = 1
ERROR = 2

log_name = ['Info:    ', 'OK:      ', 'WARNING: ', 'ERROR:   ']
file_log_level = INFO
console_log_level = INFO
log_records = []


def log(msg, error_state=-1, nocr=False):
    if error_state == -1:
        if '\u001b' in msg:
            if ESC_RED in msg:
                error_state = ERROR
            elif ESC_GREEN in msg:
                error_state = SUCCESS
            elif ESC_YELLOW in msg:
                error_state = WARNING
    else:
        if '\u001b' not in msg:
            if error_state == SUCCESS:
                msg = ESC_GREEN + msg
            elif error_state == WARNING:
                msg = ESC_YELLOW + msg
            elif error_state == ERROR:
                msg = ESC_RED + msg
    msg += ESC_RESET
    if error_state >= console_log_level or error_state == SEPARATOR:
        print(msg, flush=True)
    if error_state >= file_log_level or error_state == SEPARATOR:
        cr = "\n" if not nocr else ""

        if error_state == SEPARATOR:
            log_records.append("" + cr)
            log_records.append(kioskstdlib.erase_esc_seq(msg) + cr)
        else:
            log_records.append(kioskstdlib.erase_esc_seq(log_name[error_state + 1] + msg) + cr)


def interpret_param(known_param, param: str):
    new_option = params[known_param]

    if new_option == "db":
        rc = {new_option: param.split("=")[1]}
    elif new_option == "logfile":
        rc = {new_option: param.split("=")[1]}
    elif new_option == "user":
        rc = {new_option: param.split("=")[1]}
    else:
        rc = {new_option: None}

    return rc


def usage(error: str = ""):
    if error:
        print("\n\u001b[31;1m")
        print("    " + error + "\u001b[0m")
    print("""
    Usage of buapcmimport.py:
    ===================
      This works only with a kiosk that has "buap" as its project-id and "buap" as its database name 
      (the latter can be changed with --db)

      buapcmimport <kiosk-dir> <path (and filename of excel file) to import> [options]      

      <kiosk-dir> is required and must point to the base directory 
                  (in which either a config\kiosk_config.yml or a config\sync_config.yml is expected)

      <path and filename of Excel file to import> is required and must point to an Excel file that meets the specs. 

      options:
        --user=<user>: Obligatory! The recording user to be used when creating ceramic entries, loci etc. E.g. "lkh"
        -db=<databasename>: confirms to use the stated "databasename" as database. This database name 
                          and the database name configured in the kiosk config MUST match. 
                          Otherwise RLIIM is used as default

        --import/-i: import and analyze (if not set the tool will only analyze but not import)
        --no_context_check/-c: Does not check if the context is a KNOWN context in the database
        --log_file/-lf=<path and filename of logfile>: writes important messages to a logfile
        --dev/-d: results in additional information useful rather during development 
        --no_commit/-nc: will not commit changes if in -i mode
        --apply_imports: tries to apply the imported records to the current collected material (that's what synchronization usually does)
        --erase_former_import: will delete all collected materials that stem from a former input and ALL SMALL FINDS no matter where they come from!

    """)
    sys.exit(0)


def startup():
    global cfg, workbook_file, kiosk_dir, options, console_log_level, file_log_level
    logging.basicConfig(format='>[%(module)s.%(levelname)s at %(asctime)s]: %(message)s', level=logging.ERROR)
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    if len(sys.argv) < 2:
        usage()

    kiosk_dir = sys.argv[1]
    if kiosk_dir and not path.isdir(kiosk_dir):
        usage(f"kiosk directory {kiosk_dir} does not seem to exist or is not a valid directory.")
    workbook_file = sys.argv[2]

    if workbook_file and not path.exists(workbook_file):
        usage(f"The path or Excel file {workbook_file} does not exist.")

    cfg_file = path.join(kiosk_dir, r'config\kiosk_config.yml')
    if not path.isfile(cfg_file):
        cfg_file = path.join(kiosk_dir, r'config\sync_config.yml')
        if not path.isfile(cfg_file):
            usage(f"neither a sync_config.yml nor a kiosk_config.yml can be found. ")

    cfg = SyncConfig.get_config({'config_file': cfg_file})
    if cfg.get_project_id().lower() != "rliim" and cfg.get_project_id().lower() != "rliim_test":
        usage(f"This is not a RLIIM Kiosk!")

    for i in range(3, len(sys.argv)):
        param = sys.argv[i]
        known_param = [p for p in params if param.lower().startswith(p)]
        if known_param:
            known_param = known_param[0]
            new_option = interpret_param(known_param, param)
            if new_option:
                options.update(new_option)
            else:
                logging.error(f"parameter {param} not understood.")
                usage()
        else:
            log(f"parameter \"{param}\" unknown.")
            usage()
    if "db" in options:
        if cfg.database_name != options["db"]:
            usage(f"configured database does not match '{options['db']}! (It's '{cfg.database_name}')'")
        cfg.database_name = options["db"]

    else:
        if cfg.database_name.lower() != "rliim":
            usage(f"configured database must be 'rliim', which is not the case (It's '{cfg.database_name}')!")
            usage()
    if "user" not in options:
        usage("Please state a recording user using the --user option.")
    if "dev" not in options:
        warnings.simplefilter("ignore")
    else:
        console_log_level = -1
        file_log_level = -1


def write_log_file(filename: str):
    with open(filename, "w") as f:
        f.writelines(log_records)
    log_records.clear()


def open_file(workbook_file: str):
    wb = None
    try:
        wb = load_workbook(filename=workbook_file, read_only=True, data_only=True)
        if not wb:
            raise Exception("File was not opened")

        log(f"\u001b[32;1mUsing input file {workbook_file}\u001b[0m")
    except BaseException as e:
        logging.error(f"ustpceramicsimport: Error opening {workbook_file}: {repr(e)}")
        usage(f"ustpceramicsimport: Error opening {workbook_file}: {repr(e)}")

    return wb


# *******************************************************************************+
# import code
# *******************************************************************************+#



def check_structure(ws: Worksheet, expected_cols):
    err = False
    for col_idx, col_name in enumerate(expected_cols):
        try:
            v = ws.cell(1, col_idx + 1).value.lower()
            if v != col_name.lower():
                log(f"row 1, col ${col_idx + 1}: expected ${col_name}, got ${v}",
                    error_state=ERROR)
                err = True
        except BaseException as e:
            log(f"row 1, col ${col_idx + 1}: {repr(e)}")
            err = True

    return not err


def get_grams(gram_expr: str) -> float:
    if not gram_expr or not gram_expr.strip():
        raise ValueError("empty")

    regex = re.compile(r"^\s*(?P<num>[\d|.|,]*)\s*(?P<unit>[A-z]*)\s*$")
    match = regex.match(gram_expr)
    if match:
        num = 0.0
        unit = "g"
        groups = match.groupdict()
        if "num" in groups and groups["num"]:
            num = float(match.groupdict()["num"])
        else:
            raise ValueError("Missing numerical part")

        if "unit" in groups and groups["unit"]:
            unit = match.groupdict()["unit"].lower()

        if unit in ["g", "grams", "gr"]:
            return num
        if unit in ["kg", "k"]:
            return num * 1000
        raise ValueError("Unit unknown")
    else:
        raise ValueError("unclear weight")


def get_sf_type_from_description(description):
    regex_and_type = [("coin", "coin"),
                      ("\\bostr", "ostracon"),
                      ("falcon mummy", "falcon mummy"),
                      ("miniature vessel", "miniature vessel"),
                      ("vessel", "vessel"),
                      ("pot for ibises", "vessel"),
                      ("ibis jar", "ibis jar"),
                      ("ibis", "ibis"),
                      ("jar seal", "jar sealing"),
                      ("seal impression", "seal impression"),
                      ("rifle", "weapon"),
                      ("scarab impression", "seal impression"),
                      ("stopper", "jar stopper"),
                      ("\\bjar\\b", "vessel"),
                      ("\\bsealing\\b", "sealing"),
                      ("shabti", "ushabti"),
                      ("shawabti", "ushabti"),
                      ("amulet", "amulet"),
                      ("lithic", "lithic"),
                      ("flint", "lithic"),
                      ("chert", "lithic"),
                      ("copper.*blade", "copper blade"),
                      ("blade.*copper", "copper blade"),
                      ("blade", "lithic"),
                      ("bracelet", "bracelet"),
                      ("necklace", "necklace"),
                      ("pendant", "necklace"),
                      ("\\bstela\\b", "stela"),
                      ("\\binscri", "inscription"),
                      ("bullet", "bullet"),
                      ("figurine", "figurine"),
                      ("sculptu", "sculpture"),
                      ("statue", "sculpture"),
                      ("coffin", "coffin"),
                      ("scarab", "scarab"),
                      ("\\bdie\\b", "die"),
                      ("cordage", "cordage"),
                      ("\\bcup\\b", "vessel"),
                      ("bowl", "vessel"),
                      ("dish", "vessel"),
                      ("plate", "vessel"),
                      ("\\bpot\\b", "vessel"),
                      ("\\blids?\\b", "vessel"),
                      ("\\blamps?\\b", "vessel"),
                      ("vase", "vessel"),
                      ("earring", "earring"),
                      ("palette", "palette"),
                      ("\\bcloth\\b", "cloth"),
                      ("\\bbeads?\\b", "bead"),
                      ]
    for regex, sf_type in regex_and_type:
        if re.search(regex, description, flags=re.IGNORECASE):
            return sf_type

    return None


def map_material(material: str) -> str:
    if not material:
        return material

    map_material = material.strip().lower()
    mapping = {
        "ch stone": "chipped stone",
        "clay": "clay/mud",
        "gr stone": "ground stone",
        "egyptian blue": "Egyptian blue",
    }
    if map_material in mapping:
        return mapping[map_material]

    return material.strip()


def import_lot_rows(ws: Worksheet, max_rows=-1):
    def get_or(index, alternative):
        if index < max_col:
            if row[index].value:
                return row[index].value
        return alternative

    row_nr = 1
    rows = ws.iter_rows(min_row=2)
    to_delete = []
    successful_rows = 0
    for row in rows:
        row_nr += 1
        if row_nr % 100 == 0:
            print(".", end="", flush=True)
        if row_nr % 5000 == 0:
            print("\n", end="", flush=True)
        if max_rows > -1 and row_nr == max_rows + 2:
            print(f"Reached maximum row {row_nr}: done.")
            break

        if not row[0].value or not str(row[0].value).strip():
            print(f"Found empty first cell in row {row_nr}: done.")
            break

        max_col = len(row)
        try:
            lot_nr = int(row[0].value)
        except BaseException as e:
            log(f"Line {row_nr}: invalid Lot Number {row[0].value}. Line skipped.", error_state=WARNING)
            continue
        if lot_nr <= 0:
            log(f"Line {row_nr}: invalid Lot Number {row[0].value}. Line skipped.", error_state=WARNING)
            continue

        try:
            context = row[1].value.strip()
            if not re.search(r"^[A|B|C|D]\d\d\d$", context):
                log(f"Line {row_nr}, lot# {lot_nr}: context {context} does not look right. Line skipped.", error_state=WARNING)
                continue
        except BaseException as e:
            log(f"Line {row_nr}, lot# {lot_nr}: no context. Line skipped.", error_state=WARNING)
            continue
        trench = row[2].value.strip()
        if not re.search(r"^[A|B|C|D]$", trench):
            log(f"Line {row_nr}, lot# {lot_nr}: trench {trench} does not look right. Line skipped.", error_state=WARNING)
            continue
        else:
            if context and trench != context[0]:
                log(f"Line {row_nr}, lot# {lot_nr}: trench {trench} does not match context.", error_state=WARNING)

        lot_type = row[3].value
        if lot_type:
            lot_type = lot_type.lower()
        else:
            log(f"Line {row_nr}, LOT# {lot_nr}: type is empty.", error_state=WARNING)

        date_excavated = get_or(4, None)
        if date_excavated:
            if not isinstance(date_excavated, datetime.datetime):
                log(f"Line {row_nr}, LOT# {lot_nr}:: date excavated {date_excavated} is not a date. Value dropped.", error_state=INFO)
                date_excavated = None
        else:
            log(f"Line {row_nr}, LOT# {lot_nr}:: Date excavated is empty.", error_state=WARNING)

        description = get_or(5, None)

        date_recorded = get_or(6, None)
        if date_recorded:
            if not isinstance(date_recorded, datetime.datetime):
                log(f"Line {row_nr}, LOT# {lot_nr}:: date recorded {date_recorded} is not a date. Value dropped.", error_state=INFO)
                date_recorded = None

        # else:
        #     log(f"Line {row_nr}, LOT# {lot_nr}:: Date recorded is empty.", error_state=INFO)

        sherd_count = get_or(7, None)
        if sherd_count:
            if not isinstance(sherd_count, int) and not isinstance(sherd_count, float):
                log(f"Line {row_nr}, LOT# {lot_nr}:: sherd count {sherd_count} is not a number. Value dropped.",
                    error_state=INFO)
                sherd_count = None
            else:
                if lot_type.lower() != "ceramic":
                    log(f"Line {row_nr}, LOT# {lot_nr}:: sherd count for a collected material "
                        f"that is not of type 'ceramic'.", error_state=INFO)

        photographed = True if get_or(7, None) else False

        cm_type = "bulk"
        context = f"{context[0]}-{context[1:]}"
        arch_context = f"{context}-{lot_nr}"
        arch_domain = None
        local_nr = lot_nr

        sql = f"""
            {'INSERT'} INTO "{RLIIM_IMPORT_TABLE}" VALUES(
                %s, /* line_ref */
                %s, /* cm_type varchar, */
                %s, /* \"lot\" VARCHAR PRIMARY KEY NOT NULL, */
                %s, /* locus VARCHAR NOT NULL, */
                %s, /* unit VARCHAR NOT NULL, */
                %s, /* arch_domain VARCHAR NOT NULL, */
                %s, /* arch_context VARCHAR NOT NULL, */
                %s, /* local_nr numeric, */
                %s, /* type VARCHAR, */
                %s, /* date_excavated TIMESTAMP, */
                %s, /* date_entered TIMESTAMP, */
                %s, /* photographed BOOLEAN, */
                %s, /* description VARCHAR, */
                %s /* count NUMERIC */
            )
        """
        params = [
            row_nr,
            cm_type, lot_nr, context, trench, arch_domain, arch_context, local_nr, lot_type, date_excavated, date_recorded,
            photographed, description, sherd_count]
        try:
            KioskSQLDb.execute(sql, parameters=params, commit=True)
            successful_rows += 1
        except BaseException as e:
            if "pkey" in repr(e):
                log(f"Line {row_nr}, LOT# {lot_nr}: Duplicate lot number. "
                    f"All lines with this lot number will be skipped.", error_state=ERROR)
                KioskSQLDb.rollback()
                to_delete.append(lot_nr)
                continue
            else:
                log(f"Line {row_nr}, LOT# {lot_nr}: SQL Error '{e}'. Aborting.", error_state=ERROR)
                print(params)
                return False

    for lot in to_delete:
        c = KioskSQLDb.execute(f"delete from {RLIIM_IMPORT_TABLE} where lot=%s", parameters=[lot], commit=True)
        successful_rows -= c
    if to_delete:
        log(f"Deleted {len(to_delete)} lot numbers because they existed more than once in the source.")
    log(f"\nSuccessfully imported {successful_rows} of {row_nr} rows from sheet \"Lots\"")
    return True

def import_sample_rows(ws: Worksheet, max_rows=-1):
    def get_or(index, alternative):
        if index < max_col:
            if row[index].value:
                return row[index].value
        return alternative

    row_nr = 1
    rows = ws.iter_rows(min_row=2)
    to_delete = []
    successful_rows = 0
    for row in rows:
        row_nr += 1
        if row_nr % 100 == 0:
            print(".", end="", flush=True)
        if row_nr % 5000 == 0:
            print("\n", end="", flush=True)
        if max_rows > -1 and row_nr == max_rows + 2:
            print(f"Reached maximum row {row_nr}: done.")
            break

        if not row[0].value or not str(row[0].value).strip():
            print(f"Found empty first cell in row {row_nr}: done.")
            break

        max_col = len(row)
        try:
            lot_nr = row[0].value
            local_nr = int(lot_nr[1:])
        except BaseException as e:
            log(f"Line {row_nr}: invalid Lot Number {row[0].value}. Line skipped.", error_state=WARNING)
            continue

        if local_nr <= 0:
            log(f"Line {row_nr}: invalid Lot Number {row[0].value}. Line skipped.", error_state=WARNING)
            continue

        try:
            context = row[1].value.strip()
            if not re.search(r"^[A|B|C|D]\d\d\d$", context):
                log(f"Line {row_nr}, lot# {lot_nr}: context {context} does not look right. Line skipped.", error_state=WARNING)
                continue
        except BaseException as e:
            log(f"Line {row_nr}, lot# {lot_nr}: no context. Line skipped.", error_state=WARNING)
            continue
        trench = row[2].value.strip()
        if not re.search(r"^[A|B|C|D]$", trench):
            log(f"Line {row_nr}, lot# {lot_nr}: trench {trench} does not look right. Line skipped.", error_state=WARNING)
            continue
        else:
            if context and trench != context[0]:
                log(f"Line {row_nr}, lot# {lot_nr}: trench {trench} does not match context.", error_state=WARNING)

        lot_type = row[3].value
        if lot_type:
            lot_type = lot_type.lower()
        else:
            log(f"Line {row_nr}, LOT# {lot_nr}: type is empty.", error_state=WARNING)

        description = get_or(4, None)

        date_excavated = get_or(5, None)
        if date_excavated:
            if not isinstance(date_excavated, datetime.datetime):
                log(f"Line {row_nr}, LOT# {lot_nr}:: date excavated {date_excavated} is not a date. Value dropped.", error_state=INFO)
                date_excavated = None
        else:
            log(f"Line {row_nr}, LOT# {lot_nr}:: Date excavated is empty.", error_state=WARNING)


        date_recorded = date_excavated

        notes = get_or(6, None)
        if notes:
            description = f"{(description + ' ') if description is not None else ''}{notes}"

        location = get_or(7, None)

        cm_type = "sample"
        context = f"{context[0]}-{context[1:]}"
        arch_context = f"{context}-{lot_nr}"
        arch_domain = "S"

        sql = f"""
            {'INSERT'} INTO "{RLIIM_IMPORT_TABLE}" VALUES(
                %s, /* line_ref */
                %s, /* cm_type varchar, */
                %s, /* \"lot\" VARCHAR PRIMARY KEY NOT NULL, */
                %s, /* locus VARCHAR NOT NULL, */
                %s, /* unit VARCHAR NOT NULL, */
                %s, /* arch_domain VARCHAR NOT NULL, */
                %s, /* arch_context VARCHAR NOT NULL, */
                %s, /* local_nr numeric, */
                %s, /* type VARCHAR, */
                %s, /* date_excavated TIMESTAMP, */
                %s, /* date_entered TIMESTAMP, */
                %s, /* photographed BOOLEAN, */
                %s, /* description VARCHAR, */
                %s, /* count numeric, */
                %s /* location VARCHAR */
            )
        """
        params = [
            row_nr,
            cm_type, lot_nr, context, trench, arch_domain, arch_context, local_nr, lot_type, date_excavated,
            date_recorded, False, description, None, location]
        try:
            KioskSQLDb.execute(sql, parameters=params, commit=True)
            successful_rows += 1
        except BaseException as e:
            if "pkey" in repr(e):
                log(f"Line {row_nr}, LOT# {lot_nr}: Duplicate lot number. "
                    f"All lines with this lot number will be skipped.", error_state=ERROR)
                KioskSQLDb.rollback()
                to_delete.append(lot_nr)
                continue
            else:
                log(f"Line {row_nr}, LOT# {lot_nr}: SQL Error '{e}'. Aborting.", error_state=ERROR)
                print(params)
                return False

    for lot in to_delete:
        c = KioskSQLDb.execute(f"delete from {RLIIM_IMPORT_TABLE} where lot=%s", parameters=[lot], commit=True)
        successful_rows -= c
    if to_delete:
        log(f"Deleted {len(to_delete)} lot numbers because they existed more than once in the source.")
    log(f"\nSuccessfully imported {successful_rows} of {row_nr} rows from sheet \"Samples\"")
    return True

def import_artifacts_rows(ws: Worksheet, max_rows=-1):
    def interpret_measure_token(t):
        try:
            parts: List[str] = t.split(":")
            if len(parts) != 2:
                return "error", f": missing in {t}"
            key = parts[0].strip().lower()
            value = parts[1].strip().lower()

            if key in ["w", "l", "h"]:
                pass
            elif "perf" in key:
                key = "perf"
            elif "weight" in key:
                key = "weight"
            elif "dia" in key:
                key = "dia"
            elif "len" in key:
                key = "l"
            elif "width" in key:
                key = "w"
            elif "height" in key:
                key = "h"
            elif "thick" in key:
                key = "thickness"

            else:
                return "error", f"unclear dimension {key}"

            if "mm" in value or "cm" in value or " m" in value:
                pass
            else:
                if key == "weight" and "g" in value:
                    pass
                else:
                    return "error", f"no known measurement unit in {key}."

            if kioskstdlib.force_positive_int_from_string(t) == -1:
                return "error", f"no number in {t}"

            return key, value
        except BaseException as e:
            return "error", repr(e)

    def get_or(index, alternative):
        if index < max_col:
            if row[index].value:
                return row[index].value
        return alternative

    row_nr = 1
    rows = ws.iter_rows(min_row=2)
    to_delete = []
    successful_rows = 0
    for row in rows:
        row_nr += 1
        if row_nr % 100 == 0:
            print(".", end="", flush=True)
        if row_nr % 5000 == 0:
            print("\n", end="", flush=True)
        if max_rows > -1 and row_nr == max_rows + 2:
            print(f"Reached maximum row {row_nr}: done.")
            break

        if not row[0].value or not str(row[0].value).strip():
            print(f"Found empty first cell in row {row_nr}: done.")
            break

        max_col = len(row)
        try:
            lot_nr = int(row[0].value)
        except BaseException as e:
            log(f"Line {row_nr}: invalid Lot Number {row[0].value}. Line skipped.", error_state=WARNING)
            continue
        if lot_nr <= 0:
            log(f"Line {row_nr}: invalid Lot Number {row[0].value}. Line skipped.", error_state=WARNING)
            continue

        try:
            context = row[1].value.strip()
            if not re.search(r"^[A|B|C|D]\d\d\d$", context):
                log(f"Line {row_nr}, lot# {lot_nr}: context {context} does not look right. Line skipped.", error_state=WARNING)
                continue
        except BaseException as e:
            log(f"Line {row_nr}, lot# {lot_nr}: no context. Line skipped.", error_state=WARNING)
            continue
        trench = row[2].value.strip()
        if not re.search(r"^[A|B|C|D]$", trench):
            log(f"Line {row_nr}, lot# {lot_nr}: trench {trench} does not look right. Line skipped.", error_state=WARNING)
            continue
        else:
            if context and trench != context[0]:
                log(f"Line {row_nr}, lot# {lot_nr}: trench {trench} does not match context.", error_state=WARNING)

        sf_type = row[3].value
        if sf_type:
            sf_type = sf_type.lower()
        else:
            log(f"Line {row_nr}, LOT# {lot_nr}: type is empty.", error_state=WARNING)

        date_excavated = get_or(4, None)
        if date_excavated:
            if not isinstance(date_excavated, datetime.datetime):
                log(f"Line {row_nr}, LOT# {lot_nr}:: date excavated {date_excavated} is not a date. Value dropped.", error_state=INFO)
                date_excavated = None
        else:
            log(f"Line {row_nr}, LOT# {lot_nr}:: Date excavated is empty.", error_state=WARNING)

        measures:str = get_or(5, None)
        weight_raw = get_or(6, None)

        description = get_or(7, None)

        date_recorded = get_or(8, None)
        if date_recorded:
            if not isinstance(date_recorded, datetime.datetime):
                log(f"Line {row_nr}, LOT# {lot_nr}:: date recorded {date_recorded} is not a date. Value dropped.", error_state=INFO)
                date_recorded = None

        photographed = True if get_or(9, None) else False

        length=None
        height=None
        width=None
        thickness=None
        diameter=None
        diameter_perf=None
        weight=None

        if measures:
            measures = measures.replace(",", ";")
            token = measures.split(";")
            for t in token:
                key, value = interpret_measure_token(t)
                if key == "error":
                    log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has part \"{t}\" that I can't interpret ({value}). "
                        f"That part is dropped.",
                        error_state=INFO)
                    continue
                else:
                    if key == "w":
                        if width:
                            log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has \"{key}\" twice or more. "
                                f"That part is dropped.",
                                error_state=INFO)
                            continue
                        else:
                            width = value
                    elif key == "l":
                        if length:
                            log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has \"{key}\" twice or more. "
                                f"That part is dropped.",
                                error_state=INFO)
                            continue
                        else:
                            length = value
                    elif key == "h":
                        if height:
                            log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has \"{key}\" twice or more. "
                                f"That part is dropped.",
                                error_state=INFO)
                            continue
                        else:
                            height = value
                    elif key == "thickness":
                        if thickness:
                            log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has \"{key}\" twice or more. "
                                f"That part is dropped.",
                                error_state=INFO)
                            continue
                        else:
                            thickness = value
                    elif key == "dia":
                        if diameter:
                            log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has \"{key}\" twice or more. "
                                f"That part is dropped.",
                                error_state=INFO)
                            continue
                        else:
                            diameter = value
                    elif key == "perf":
                        if diameter_perf:
                            log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has \"{key}\" twice or more. "
                                f"That part is dropped.",
                                error_state=INFO)
                            continue
                        else:
                            diameter_perf = value
                    else:
                        log(f"Line {row_nr}, LOT# {lot_nr}:: Other Dimensions has \"{key}\". What is that? "
                            f"That part is dropped.",
                            error_state=ERROR)
                        continue
        if weight_raw:
            key, value = interpret_measure_token("weight: " + weight_raw)
            if key == "error":
                log(f"Line {row_nr}, LOT# {lot_nr}:: Weight \"{weight_raw}\" looks fishy ({value}). "
                    f"Value dropped.",
                    error_state=INFO)
                continue
            else:
                weight = kioskstdlib.force_positive_int_from_string(weight_raw)

        cm_type = "small_find"
        context = f"{context[0]}-{context[1:]}"
        c = KioskSQLDb.get_record_count(RLIIM_IMPORT_TABLE, "line_ref", "lot = %s", [str(lot_nr)])
        if c == 0:
            log(f"Line {row_nr}, LOT# {lot_nr}:: There is no record in Lots that "
                f"matches line {row_nr} in Artifacts. line skipped!",
                error_state=ERROR)
            continue
        if c > 1:
            log(f"Line {row_nr}, LOT# {lot_nr}:: There is more than one record in Lots that "
                f"matches line {row_nr} in Artifacts. That should not be the case at all! line skipped!",
                error_state=ERROR)
            continue
        r = KioskSQLDb.get_first_record(RLIIM_IMPORT_TABLE, "lot", str(lot_nr))
        if context != r["locus"]:
            log(f"Line {row_nr}, LOT# {lot_nr}:: The context in artifacts does not match the context in Lots. line skipped!",
                error_state=ERROR)
            continue

        sql = f"""
            UPDATE {RLIIM_IMPORT_TABLE}
            SET line_ref=%s,
                cm_type=%s,
                sf_type=%s,
                sf_measures=%s,
                sf_length_mm=%s,
                sf_height_mm=%s,
                sf_width_mm=%s,
                sf_thickness_mm=%s,
                sf_diameter_mm=%s,
                sf_diameter_perf_mm=%s,
                sf_weight=%s,
                sf_description=%s,
                sf_date_excavated=%s,
                sf_date_entered=%s,
                sf_photographed=%s
           WHERE lot=%s     
        """
        params = [
            row_nr,
            cm_type, sf_type, measures, length, height, width, thickness, diameter, diameter_perf, weight, description,
            date_excavated, date_recorded, photographed, str(lot_nr)]
        try:
            KioskSQLDb.execute(sql, parameters=params, commit=True)
            successful_rows += 1
        except BaseException as e:
            log(f"Line {row_nr}, LOT# {lot_nr}: SQL Error '{e}'. Aborting.", error_state=ERROR)
            print(params)
            return False

    log(f"\nSuccessfully imported {successful_rows} of {row_nr} rows from sheet \"Artifacts\"")
    return True

def create_import_table():
    KioskSQLDb.drop_table_if_exists(RLIIM_IMPORT_TABLE)
    sql = f"""
        {'create'} table \"{RLIIM_IMPORT_TABLE}\"(
            line_ref VARCHAR NOT NULL,
            cm_type varchar NOT NULL,
            \"lot\" VARCHAR PRIMARY KEY NOT NULL,
            locus VARCHAR NOT NULL,
            unit VARCHAR NOT NULL,
            arch_domain VARCHAR,
            arch_context VARCHAR NOT NULL,
            local_nr numeric NOT NULL,
            \"type\" VARCHAR,
            date_excavated TIMESTAMP,
            date_entered TIMESTAMP,
            photographed BOOLEAN,
            description VARCHAR,
            \"count\" NUMERIC,
            location VARCHAR,
            sf_type VARCHAR,
            sf_measures VARCHAR,
            sf_length_mm VARCHAR,
            sf_height_mm VARCHAR,
            sf_width_mm VARCHAR,
            sf_thickness_mm VARCHAR,
            sf_diameter_mm VARCHAR,
            sf_diameter_perf_mm VARCHAR,
            sf_weight NUMERIC,
            sf_description VARCHAR,
            sf_date_excavated TIMESTAMP,
            sf_date_entered TIMESTAMP,
            sf_photographed BOOLEAN
            )"""

    KioskSQLDb.execute(sql, commit=True)

def import_workbook():
    wb: Workbook = open_file(workbook_file)
    try:
        rc = 0
        log(f"*" * 80, error_state=SEPARATOR)
        log(f" Workbook '{kioskstdlib.get_filename(workbook_file)}'  ".center(80, "*"), error_state=SEPARATOR,
            nocr=True)
        log(f"*" * 80, error_state=SEPARATOR)

        sheet_structures = {}
        sheet_structures["Lots"] = expected_cols_lots
        sheet_structures["Samples"] = expected_cols_samples
        sheet_structures["Artifacts"] = expected_cols_artifacts

        for sheet in ["Lots", "Samples", "Artifacts"]:
            if sheet not in wb.sheetnames:
                log(f"there is no sheet \"{sheet}\" in this workbook.", error_state=ERROR)

            ws: Worksheet = wb[sheet]
            if not check_structure(ws, sheet_structures[sheet]):
                log(f"there are structural errors in the worksheet '{sheet}': Aborting.", error_state=ERROR)
            else:
                log(f"Worksheet structure for sheet {sheet} okay.", error_state=INFO)

        log(f"--------------------------------------------", error_state=INFO)

        log(f"Importing Worksheet Lots:", error_state=INFO)
        ws: Worksheet = wb["Lots"]
        import_lot_rows(ws)

        log(f"Importing Worksheet Samples:", error_state=INFO)
        ws: Worksheet = wb["Samples"]
        import_sample_rows(ws)

        log(f"Importing Worksheet Artifacts:", error_state=INFO)
        ws: Worksheet = wb["Artifacts"]
        import_artifacts_rows(ws)
    finally:
        wb.close()


def erase_former_import():
    try:
        KioskSQLDb.execute("""
        delete from small_find finds where uid_cm in(
            select uid from collected_material where collected_material.dearregistrar = 'imported by rliim kiosk bridge')
        """)
        KioskSQLDb.execute("""
        delete from collected_material_photo where uid_cm in(
            select uid from collected_material where collected_material.dearregistrar = 'imported by rliim kiosk bridge')
        """)
        KioskSQLDb.execute("""
        delete from collected_material where  collected_material.dearregistrar = 'imported by rliim kiosk bridge'
        """)
        KioskSQLDb.commit()
        log("Successfully erased former import.", error_state=INFO)
        return True
    except BaseException as e:
        log(f"ERROR IN ease_former_input: {repr(e)}", error_state=ERROR)
        try:
            KioskSQLDb.rollback()
        except BaseException as e:
            pass

    return False


# *******************************************************************************+
# MAIN
# *******************************************************************************+#

if __name__ == '__main__':

    startup()

    if "logfile" in options:
        log_path = kioskstdlib.get_file_path(options["logfile"])
        log_file = options["logfile"]
    else:
        log_file = "rliimcmimport.log"
        log_path = kioskstdlib.get_file_path(workbook_file)

    print("", flush=True)
    print("\u001b[2J", flush=True)
    log("************************************************************************", error_state=0)
    log("******         \u001b[34mRLIIM Collected Material Import Tool V1.0 \u001b[0m         ******", error_state=0)
    log("************************************************************************", error_state=0)
    log(f"\u001b[32;1mOperating on database {cfg.database_name}\u001b[0m", error_state=0)
    log(f"\u001b[32;1mImporting file {workbook_file}\u001b[0m", error_state=0)
    log(f"\u001b[32;1mWith options {options}\u001b[0m", error_state=0)
    print(f"")
    create_import_table()
    import_workbook()
    # KioskSQLDb.execute("update collected_material set external_id = null", commit=True)
    if "efi" in options:
        if not erase_former_import():
            exit(-1)
    if "apply" in options:
        import_rliim_cm_for_new_loci()
    write_log_file(path.join(log_path, kioskstdlib.get_filename_without_extension(log_file) + ".log"))
